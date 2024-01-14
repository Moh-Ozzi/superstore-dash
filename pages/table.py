import dash
import dash_ag_grid as dag
import dash_bootstrap_components as dbc
from dash import html, dcc, Output, Input, callback, State, ctx
from flask_login import current_user
from utils.login_handler import require_login
from pages.funcs import create_main_df
import plotly.express as px
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
import os
import shutil
from io import BytesIO




dash.register_page(__name__, title='table', path='/table')
require_login(__name__)


df = create_main_df()
df = df[['order_id', 'order_date', 'product_name', 'customer_name', 'ship_mode', 'state', 'category', 'quantity', 'sales', 'profit']]
df['sales'] = df['sales'].round(2)
df['profit'] = df['profit'].round(2)


cellStyle = {
    "styleConditions": [
        {
            "condition": "params.data.profit < 0",
            "style": {"backgroundColor": "#FD1C03"},
        },
    ]
}

# columns_def = [
#     {"field": col, "filter": "agNumberColumnFilter", 'type': 'rightAligned'} if str(df[col].dtype) in ['double[pyarrow]', 'int64[pyarrow]'] else {"field": col, "filter": "agSetColumnFilter"}
#     for col in df.columns
# ]

columns_def = [
    {"field": "order_id", "filter": "agSetColumnFilter", "cellStyle": cellStyle},
    {"field": "order_date"},
    {"field": "product_name", "filter": "agSetColumnFilter"},
    {"field": "customer_name", "filter": "agSetColumnFilter"},
    {"field": "state", "filter": "agSetColumnFilter"},
    {"field": "ship_mode", "filter": "agSetColumnFilter"},
    {"field": "category", "filter": "agSetColumnFilter"},
    {"field": "quantity", "filter": "agNumberColumnFilter", 'type': 'rightAligned'},
    {"field": "sales", "filter": "agNumberColumnFilter", 'type': 'rightAligned'},
    {"field": "profit", "filter": "agNumberColumnFilter", 'type': 'rightAligned', "cellStyle": cellStyle},
]


table = dag.AgGrid(id='table',
                   columnDefs=columns_def,
                   rowData=df.to_dict('records'),
                   defaultColDef={"resizable": True, "sortable": True, 'editable': True, "filter": True},
                   columnSize="sizeToFit",
                   enableEnterpriseModules=True,
                   dashGridOptions={"pagination": True, "paginationPageSize": 20},
                   className="ag-theme-alpine",
style={"height": 600, "width": '100%'}

                   )
#
# htc_revenue = df['htc_revenue'].sum() / 1000
layout = dbc.Container(
    [
        html.Div(dbc.Button('Export', id='button', className="m-2")),
        table,
        html.H4(id='text_holder', children=''),
        dcc.Download(id="download"),
    ]
    )


@callback(Output("download", "data"), Input('button', 'n_clicks'), Input("table", "virtualRowData"))
def export_data(clicks, vdata):
    if clicks and vdata:
        dff2 = pd.DataFrame(vdata)
        dff2.drop(columns=['product_name'], inplace=True)
        dff2 = [dff2.columns.tolist()] + dff2.values.tolist()


        # Create a new workbook and sheet
        wb = openpyxl.Workbook()
        sheet = wb.active

        # Populate the sheet with data
        for row in dff2:
            sheet.append(row)

        # Apply formatting to the header row
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border

        # Apply conditional formatting to the "Age" column
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"))

        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=9):
            cell_col9 = row[8]
            cell_col1 = row[0]
            if cell_col9.value is not None and cell_col9.value < 0:
                cell_col9.fill = red_fill
                cell_col9.font = Font(color="FFFFFF")
                cell_col1.fill = red_fill
                cell_col1.font = Font(color="FFFFFF")
            cell_col9.border = border
            cell_col1.border = border

        # Compute the sum of the "Age" column
        # quantity_sum = sum(cell.value for cell in sheet["H"][1:])
        # sales_sum = sum(cell.value for cell in sheet["I"][1:])
        # profit_sum = sum(cell.value for cell in sheet["J"][1:])

        # Add a row for the total
        total_row = [None, None, None, None, None, None, f"=SUBTOTAL(109, G2:G{sheet.max_row})", f"=SUBTOTAL(109, H2:H{sheet.max_row})", f"=SUBTOTAL(109, I2:I{sheet.max_row})"]
        sheet.append(total_row)

        # Apply formatting to the total row
        total_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
        total_font = Font(bold=True, color="17202A")
        total_border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"))

        # Get the last row index (the total row) using the number of rows in the sheet
        last_row_index = sheet.max_row

        for cell in sheet[last_row_index]:
            cell.fill = total_fill
            cell.font = total_font
            cell.border = total_border
        # Create Excel table
        table = openpyxl.worksheet.table.Table(displayName="MyTable", ref=f"A1:I{last_row_index}")
        table.tableStyleInfo = openpyxl.worksheet.table.TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                                                       showLastColumn=False, showRowStripes=True,
                                                                       showColumnStripes=False)
        sheet.add_table(table)
        # Save the workbook with the total row
        # output_filename = "superstore.xlsx"
        # wb.save(output_filename)

        output = BytesIO()
        wb.save(output)

        # Seek to the beginning of the BytesIO stream
        output.seek(0)

        # Return the content for download
        return dcc.send_bytes(output.read(), filename="data.xlsx")

        # # Get the user's desktop path
        # user_desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        #
        # # Define the destination path on the desktop
        # destination_path = os.path.join(user_desktop, output_filename)
        #
        # # Copy the file to the desktop
        # shutil.copy(output_filename, destination_path)
        #
        # print(f"File '{output_filename}' published on your desktop.")

    #     return 'Your file has been exported on your desktop'
    # return ''